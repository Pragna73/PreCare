import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Common Color Palette
HEADER_BG = "1E7145"      # Forest Green
HEADER_BG_DARK = "0B132B" # Dark Slate
ROW_EVEN_BG = "FFFFFF"
ROW_ODD_BG = "F9FAFB"
BORDER_COLOR = "D0D5DD"

PASS_BG = "D4EFDF"
PASS_TEXT = "196F3D"
FAIL_BG = "FADBD8"
FAIL_TEXT = "922B21"
SKIP_BG = "FCF3CF"
SKIP_TEXT = "B7950B"


def create_three_sheet_excel_report(output_path, summary_dict, test_cases_list, failed_cases_list=None):
    """
    Creates a professional 3-sheet Excel report:
    - Sheet 1: Summary (Key-Value metrics)
    - Sheet 2: Test Cases (Complete tabular list of all executed cases)
    - Sheet 3: Failed Test Cases (Details of failures or clean 'No failed test cases' banner)
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb = openpyxl.Workbook()

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # ─────────────────────────────────────────────────────────────
    # SHEET 1: Summary
    # ─────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum["A1"]
    title_cell.value = summary_dict.get("Report Title", "PreCare Automated Test Execution Report")
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    # Summary Table Headers
    ws_sum.append(["Metric", "Value"])
    ws_sum.row_dimensions[2].height = 24
    for c_idx in range(1, 3):
        c = ws_sum.cell(row=2, column=c_idx)
        c.fill = PatternFill(start_color=HEADER_BG_DARK, end_color=HEADER_BG_DARK, fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="left" if c_idx == 1 else "center", vertical="center")
        c.border = thin_border

    # Summary Rows
    for r_idx, (k, v) in enumerate(summary_dict.items(), start=3):
        if k == "Report Title":
            continue
        ws_sum.append([k, str(v)])
        ws_sum.row_dimensions[r_idx].height = 20
        c1 = ws_sum.cell(row=r_idx, column=1)
        c2 = ws_sum.cell(row=r_idx, column=2)
        c1.font = Font(name="Segoe UI", size=10, bold=True, color="1F2937")
        c2.font = Font(name="Segoe UI", size=10, color="111827")
        c1.border = thin_border
        c2.border = thin_border
        c1.alignment = Alignment(horizontal="left", vertical="center")

        # Highlight overall result
        if k in ["Overall Result", "Overall Security Status", "Status"]:
            if str(v).upper() in ["PASSED", "PASS", "SECURE", "0 VULNERABILITIES"]:
                c2.fill = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
                c2.font = Font(name="Segoe UI", size=10, bold=True, color=PASS_TEXT)
            else:
                c2.fill = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")
                c2.font = Font(name="Segoe UI", size=10, bold=True, color=FAIL_TEXT)
            c2.alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.column_dimensions["A"].width = 34
    ws_sum.column_dimensions["B"].width = 50

    # ─────────────────────────────────────────────────────────────
    # SHEET 2: Test Cases
    # ─────────────────────────────────────────────────────────────
    ws_cases = wb.create_sheet(title="Test Cases")
    ws_cases.views.sheetView[0].showGridLines = True

    case_headers = [
        "Test Case ID", "Module", "Test Type", "Test Case Title", "Description",
        "Preconditions", "Test Steps", "Expected Result", "Actual Result",
        "Status", "Duration", "Severity", "Endpoint / Screen", "Error / Exception"
    ]
    ws_cases.append(case_headers)
    ws_cases.row_dimensions[1].height = 28

    for col_idx in range(1, len(case_headers) + 1):
        c = ws_cases.cell(row=1, column=col_idx)
        c.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    ws_cases.freeze_panes = "A2"

    for r_idx, tc in enumerate(test_cases_list, start=2):
        status_val = tc.get("Status", "PASS").upper()
        row_vals = [
            tc.get("Test Case ID", f"TC-{r_idx-1:04d}"),
            tc.get("Module", "General"),
            tc.get("Test Type", "Functional"),
            tc.get("Test Case Title", ""),
            tc.get("Description", ""),
            tc.get("Preconditions", "Application running in test environment"),
            tc.get("Test Steps", "1. Execute test\n2. Verify result"),
            tc.get("Expected Result", "Expected validation criteria met"),
            tc.get("Actual Result", "Observed criteria matched successfully"),
            status_val,
            tc.get("Duration", "0.015s"),
            tc.get("Severity", "NORMAL"),
            tc.get("Endpoint / Screen", "N/A"),
            tc.get("Error / Exception", "None")
        ]
        ws_cases.append(row_vals)
        ws_cases.row_dimensions[r_idx].height = 20

        bg_color = ROW_ODD_BG if r_idx % 2 == 1 else ROW_EVEN_BG
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        for col_idx in range(1, len(case_headers) + 1):
            cell = ws_cases.cell(row=r_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=9.5)
            cell.fill = row_fill

            if col_idx in [1, 3, 10, 11, 12]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Format Status cell
            if col_idx == 10:
                if status_val == "PASS" or status_val == "PASSED":
                    cell.fill = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=9.5, bold=True, color=PASS_TEXT)
                elif status_val == "FAIL" or status_val == "FAILED":
                    cell.fill = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=9.5, bold=True, color=FAIL_TEXT)
                else:
                    cell.fill = PatternFill(start_color=SKIP_BG, end_color=SKIP_BG, fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=9.5, bold=True, color=SKIP_TEXT)

    # Auto-adjust column width
    for col in ws_cases.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws_cases.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    ws_cases.auto_filter.ref = f"A1:{get_column_letter(len(case_headers))}{len(test_cases_list)+1}"

    # ─────────────────────────────────────────────────────────────
    # SHEET 3: Failed Test Cases
    # ─────────────────────────────────────────────────────────────
    ws_failed = wb.create_sheet(title="Failed Test Cases")
    ws_failed.views.sheetView[0].showGridLines = True

    failed_headers = [
        "Test Case ID", "Module", "Test Case Title", "Description",
        "Expected Result", "Actual Result", "Failure Reason", "Error Message",
        "Severity", "Timestamp", "Recommended Action"
    ]
    ws_failed.append(failed_headers)
    ws_failed.row_dimensions[1].height = 28

    for col_idx in range(1, len(failed_headers) + 1):
        c = ws_failed.cell(row=1, column=col_idx)
        c.fill = PatternFill(start_color="922B21" if failed_cases_list else HEADER_BG, end_color="922B21" if failed_cases_list else HEADER_BG, fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    ws_failed.freeze_panes = "A2"

    if failed_cases_list and len(failed_cases_list) > 0:
        for r_idx, fc in enumerate(failed_cases_list, start=2):
            row_vals = [
                fc.get("Test Case ID", ""),
                fc.get("Module", ""),
                fc.get("Test Case Title", ""),
                fc.get("Description", ""),
                fc.get("Expected Result", ""),
                fc.get("Actual Result", ""),
                fc.get("Failure Reason", ""),
                fc.get("Error Message", ""),
                fc.get("Severity", "HIGH"),
                fc.get("Timestamp", ""),
                fc.get("Recommended Action", "")
            ]
            ws_failed.append(row_vals)
            ws_failed.row_dimensions[r_idx].height = 20
            for col_idx in range(1, len(failed_headers) + 1):
                cell = ws_failed.cell(row=r_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Segoe UI", size=9.5)
                cell.fill = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")
                if col_idx in [1, 9, 10]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_failed.auto_filter.ref = f"A1:{get_column_letter(len(failed_headers))}{len(failed_cases_list)+1}"
    else:
        # Single row indicating no failures
        ws_failed.merge_cells("A2:K2")
        no_fail_cell = ws_failed["A2"]
        no_fail_cell.value = "No failed test cases. All executed test cases passed successfully."
        no_fail_cell.font = Font(name="Segoe UI", size=10, bold=True, color=PASS_TEXT)
        no_fail_cell.fill = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
        no_fail_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_failed.row_dimensions[2].height = 26
        for col_idx in range(1, len(failed_headers) + 1):
            ws_failed.cell(row=2, column=col_idx).border = thin_border

    for col in ws_failed.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws_failed.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)

    wb.save(output_path)
    print(f"✓ Saved 3-Sheet Excel Report: {output_path}")


def create_standalone_html_report(output_path, title, summary_dict, test_cases_list):
    """Creates a styled, responsive standalone HTML test report."""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    passed_count = sum(1 for tc in test_cases_list if tc.get("Status", "").upper() in ["PASS", "PASSED"])
    failed_count = sum(1 for tc in test_cases_list if tc.get("Status", "").upper() in ["FAIL", "FAILED"])
    skipped_count = len(test_cases_list) - passed_count - failed_count
    pass_pct = f"{(passed_count / len(test_cases_list) * 100):.1f}%" if test_cases_list else "0%"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #0B132B; color: #E2E8F0; padding: 30px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .header {{ background: #1C2541; padding: 25px; border-radius: 12px; border: 1px solid #3A506B; margin-bottom: 25px; }}
        .header h1 {{ color: #00F2FE; font-size: 24px; margin-bottom: 8px; }}
        .header p {{ color: #94A3B8; font-size: 14px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 25px; }}
        .stat-card {{ background: #1C2541; padding: 18px; border-radius: 10px; border: 1px solid #3A506B; text-align: center; }}
        .stat-card .val {{ font-size: 28px; font-weight: bold; margin-bottom: 5px; }}
        .stat-card .lbl {{ font-size: 12px; color: #94A3B8; text-transform: uppercase; letter-spacing: 1px; }}
        .val-pass {{ color: #10B981; }}
        .val-fail {{ color: #EF4444; }}
        .val-cyan {{ color: #00F2FE; }}
        table {{ width: 100%; border-collapse: collapse; background: #1C2541; border-radius: 10px; overflow: hidden; border: 1px solid #3A506B; margin-top: 20px; }}
        th {{ background: #1E7145; color: white; text-align: left; padding: 12px 14px; font-size: 13px; font-weight: 600; }}
        td {{ padding: 10px 14px; border-bottom: 1px solid #283759; font-size: 12.5px; }}
        tr:hover {{ background: #233157; }}
        .badge {{ padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; }}
        .badge-pass {{ background: #064E3B; color: #6EE7B7; }}
        .badge-fail {{ background: #7F1D1D; color: #FCA5A5; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{title}</h1>
            <p>PreCare Automated Verification & Quality Assurance Suite</p>
        </div>
        <div class="stats-grid">
            <div class="stat-card"><div class="val val-cyan">{len(test_cases_list)}</div><div class="lbl">Total Tests</div></div>
            <div class="stat-card"><div class="val val-pass">{passed_count}</div><div class="lbl">Passed</div></div>
            <div class="stat-card"><div class="val val-fail">{failed_count}</div><div class="lbl">Failed</div></div>
            <div class="stat-card"><div class="val val-pass">{pass_pct}</div><div class="lbl">Pass Rate</div></div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Test ID</th>
                    <th>Module</th>
                    <th>Test Case Title</th>
                    <th>Endpoint / Target</th>
                    <th>Status</th>
                    <th>Duration</th>
                </tr>
            </thead>
            <tbody>
"""
    for tc in test_cases_list[:500]: # display up to 500 rows
        status_val = tc.get("Status", "PASS").upper()
        badge_cls = "badge-pass" if status_val in ["PASS", "PASSED"] else "badge-fail"
        html += f"""
                <tr>
                    <td><code>{tc.get('Test Case ID', '')}</code></td>
                    <td>{tc.get('Module', '')}</td>
                    <td>{tc.get('Test Case Title', '')}</td>
                    <td><code>{tc.get('Endpoint / Screen', 'N/A')}</code></td>
                    <td><span class="badge {badge_cls}">{status_val}</span></td>
                    <td>{tc.get('Duration', '0.01s')}</td>
                </tr>"""

    html += """
            </tbody>
        </table>
    </div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✓ Saved HTML Report: {output_path}")
